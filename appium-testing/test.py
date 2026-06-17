import os
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from appium import webdriver
from appium.options.common import AppiumOptions
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

# Setup file paths
APK_PATH = os.path.abspath(os.path.join(os.path.dirname(__file__), '../app/build/outputs/apk/debug/app-debug.apk'))
REPORT_FILE = os.path.join(os.path.dirname(__file__), 'Appium_Test_Report.xlsx')
APPIUM_SERVER_URL = 'http://127.0.0.1:4723'

test_results = []

def log_test_result(test_id, name, description, status, duration, error=''):
    test_results.append({
        'id': test_id,
        'name': name,
        'description': description,
        'status': status,
        'duration': duration,
        'error': error
    })
    print(f"[{status}] {test_id}: {name} ({duration}ms) {f'- {error}' if error else ''}")

def generate_excel_report():
    print("Generating Appium test report in Excel...")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "E2E Mobile Test Summary"
    ws.sheet_view.showGridLines = True

    # Title Banner
    ws.merge_cells('A1:F1')
    title_cell = ws['A1']
    title_cell.value = 'TruthGuard - Appium Android E2E Test Report'
    title_cell.font = Font(name='Outfit', size=14, bold=True, color='FFFFFF')
    title_cell.fill = PatternFill(start_color='1F3A60', end_color='1F3A60', fill_type='solid')
    title_cell.alignment = Alignment(vertical='center', horizontal='center')
    ws.row_dimensions[1].height = 40

    ws.append([]) # Row 2 spacer

    # Metadata
    ws.append(['Platform', 'Android OS (Emulator)', 'Date', time.strftime("%Y-%m-%d"), 'Engine', 'Appium + Python Client'])
    ws.append(['Target APK', 'app-debug.apk', 'Time', time.strftime("%H:%M:%S"), 'Status', 'Completed'])
    
    label_font = Font(name='Plus Jakarta Sans', size=10, bold=True)
    for row in [3, 4]:
        ws.row_dimensions[row].height = 20
        for col in ['A', 'C', 'E']:
            ws[f'{col}{row}'].font = label_font

    ws.append([]) # Row 5 spacer

    # Summary Stats
    total_tests = len(test_results)
    passed_tests = len([t for t in test_results if t['status'] == 'PASSED'])
    failed_tests = total_tests - passed_tests
    pass_rate = int((passed_tests / total_tests) * 100) if total_tests > 0 else 0

    ws.append(['Total Tests', total_tests, 'Passed', passed_tests, 'Failed', failed_tests])
    ws.append(['Pass Rate', f"{pass_rate}%"])
    
    stats_font = Font(name='Plus Jakarta Sans', bold=True)
    for row in [6, 7]:
        ws.row_dimensions[row].height = 20
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws[f'{col}{row}'].font = stats_font

    ws['A6'].fill = PatternFill(start_color='EAECEE', end_color='EAECEE', fill_type='solid')
    ws['C6'].fill = PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid')
    ws['E6'].fill = PatternFill(start_color='F8D7DA', end_color='F8D7DA', fill_type='solid')
    ws['A7'].fill = PatternFill(start_color='EAECEE', end_color='EAECEE', fill_type='solid')

    ws.append([]) # Row 8 spacer

    # Table Headers
    header_row_number = 9
    headers = ['Test ID', 'Test Case Name', 'Description', 'Status', 'Duration (ms)', 'Error / Assertion Logs']
    ws.append(headers)
    ws.row_dimensions[header_row_number].height = 28

    thin_border_side = Side(border_style='thin', color='BDC3C7')
    header_border = Border(top=thin_border_side, bottom=Side(border_style='medium', color='1F3A60'), left=thin_border_side, right=thin_border_side)
    header_font = Font(name='Plus Jakarta Sans', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='34495E', end_color='34495E', fill_type='solid')

    for col_idx, col_letter in enumerate(['A', 'B', 'C', 'D', 'E', 'F'], start=1):
        cell = ws[f'{col_letter}{header_row_number}']
        cell.font = header_font
        cell.fill = header_fill
        cell.border = header_border
        cell.alignment = Alignment(vertical='center', horizontal='center' if col_idx in [1, 4] else 'left')

    # Data Rows styling
    thin_border = Border(top=Side(style='thin', color='E0E0E0'),
                         bottom=Side(style='thin', color='E0E0E0'),
                         left=Side(style='thin', color='E0E0E0'),
                         right=Side(style='thin', color='E0E0E0'))
    
    pass_fill = PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid')
    fail_fill = PatternFill(start_color='F8D7DA', end_color='F8D7DA', fill_type='solid')
    pass_font = Font(name='Plus Jakarta Sans', bold=True, color='155724', size=10)
    fail_font = Font(name='Plus Jakarta Sans', bold=True, color='721C24', size=10)

    for idx, test in enumerate(test_results):
        row_num = header_row_number + 1 + idx
        ws.append([test['id'], test['name'], test['description'], test['status'], test['duration'], test['error']])
        ws.row_dimensions[row_num].height = 24
        
        for col_idx, col_letter in enumerate(['A', 'B', 'C', 'D', 'E', 'F'], start=1):
            cell = ws[f'{col_letter}{row_num}']
            cell.font = Font(name='Plus Jakarta Sans', size=10)
            cell.border = thin_border
            
            if col_idx == 1:
                cell.alignment = Alignment(horizontal='center')
            elif col_idx == 4:
                cell.alignment = Alignment(horizontal='center')
            elif col_idx == 5:
                cell.alignment = Alignment(horizontal='right')
                
            if col_idx == 4:
                if cell.value == 'PASSED':
                    cell.font = pass_font
                    cell.fill = pass_fill
                else:
                    cell.font = fail_font
                    cell.fill = fail_fill

    # Set Widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 28
    ws.column_dimensions['C'].width = 45
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 50

    wb.save(REPORT_FILE)
    print(f"Excel report saved successfully to: {REPORT_FILE}")

def run_appium_tests():
    print("Verifying Android package configurations...")
    if not os.path.exists(APK_PATH):
        raise FileNotFoundError(f"Android build file not found at: {APK_PATH}. Run compilation first.")

    options = AppiumOptions()
    options.set_capability('platformName', 'Android')
    options.set_capability('automationName', 'UiAutomator2')
    options.set_capability('deviceName', 'Pixel_6')
    options.set_capability('app', APK_PATH)
    options.set_capability('appPackage', 'com.samavaishnavi.truthguard')
    options.set_capability('appActivity', 'com.samavaishnavi.truthguard.MainActivity')
    options.set_capability('noReset', False)
    options.set_capability('newCommandTimeout', 300)

    driver = None
    
    # TC-001: Launch Mobile App Session
    start_time = time.time()
    try:
        print("Connecting to Appium and deploying app...")
        driver = webdriver.Remote(APPIUM_SERVER_URL, options=options)
        log_test_result('TC-001', 'Launch Android App', 'Initialize Appium session and launch the TruthGuard APK.', 'PASSED', int((time.time() - start_time) * 1000))
    except Exception as e:
        log_test_result('TC-001', 'Launch Android App', 'Initialize Appium session and launch the TruthGuard APK.', 'FAILED', int((time.time() - start_time) * 1000), str(e))
        raise e

    # TC-002: Verify Home Screen Loaded
    start_time = time.time()
    try:
        # Case-insensitive title assertion (lenient selector)
        title_el = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, '//android.widget.TextView[contains(@text, "TRUTH") or contains(@text, "Truth")]'))
        )
        log_test_result('TC-002', 'Verify Home Screen Title', 'Check if TRUTHGUARD app title exists on launch.', 'PASSED', int((time.time() - start_time) * 1000))
    except Exception as e:
        log_test_result('TC-002', 'Verify Home Screen Title', 'Check if TRUTHGUARD app title exists on launch.', 'FAILED', int((time.time() - start_time) * 1000), str(e))

    # TC-003: Navigate to Verify Screen
    start_time = time.time()
    try:
        verify_btn = WebDriverWait(driver, 8).until(
            EC.presence_of_element_located((By.XPATH, '//android.widget.TextView[@text="🔍 Verify News"]'))
        )
        verify_btn.click()
        
        # Verify text field appeared
        WebDriverWait(driver, 8).until(
            EC.presence_of_element_located((By.XPATH, '//android.widget.EditText'))
        )
        log_test_result('TC-003', 'Navigate to Verify Screen', 'Click "Verify News" button and confirm input field exists.', 'PASSED', int((time.time() - start_time) * 1000))
    except Exception as e:
        log_test_result('TC-003', 'Navigate to Verify Screen', 'Click "Verify News" button and confirm input field exists.', 'FAILED', int((time.time() - start_time) * 1000), str(e))

    # TC-004: Test Fake News Verification logic
    start_time = time.time()
    try:
        text_input = WebDriverWait(driver, 5).until(
            EC.presence_of_element_located((By.XPATH, '//android.widget.EditText'))
        )
        text_input.send_keys('This is a completely fake and clickbait shock news story.')
        
        # Click Analyze
        analyze_btn = driver.find_element(By.XPATH, '//android.widget.Button')
        analyze_btn.click()
        
        # Verify result card
        result_text = WebDriverWait(driver, 6).until(
            EC.presence_of_element_located((By.XPATH, '//android.widget.TextView[contains(@text, "Likely Fake")]'))
        )
        
        confidence_text = driver.find_element(By.XPATH, '//android.widget.TextView[contains(@text, "Confidence")]')
        confidence_val = confidence_text.text
        
        if '88%' in confidence_val:
            log_test_result('TC-004', 'Verify Fake News Analysis', 'Verify inputting fake news text returns 88% confidence Fake News.', 'PASSED', int((time.time() - start_time) * 1000))
        else:
            raise ValueError(f"Confidence value mismatch: got '{confidence_val}'")
    except Exception as e:
        log_test_result('TC-004', 'Verify Fake News Analysis', 'Verify inputting fake news text returns 88% confidence Fake News.', 'FAILED', int((time.time() - start_time) * 1000), str(e))

    # TC-005: Test Genuine News Verification logic
    start_time = time.time()
    try:
        text_input = WebDriverWait(driver, 5).until(
            EC.presence_of_element_located((By.XPATH, '//android.widget.EditText'))
        )
        text_input.clear()
        text_input.send_keys('NASA discovers new solar systems with active water and clouds.')
        
        analyze_btn = driver.find_element(By.XPATH, '//android.widget.Button')
        analyze_btn.click()
        
        result_text = WebDriverWait(driver, 6).until(
            EC.presence_of_element_located((By.XPATH, '//android.widget.TextView[contains(@text, "Likely Genuine")]'))
        )
        
        confidence_text = driver.find_element(By.XPATH, '//android.widget.TextView[contains(@text, "Confidence")]')
        confidence_val = confidence_text.text
        
        if '94%' in confidence_val:
            log_test_result('TC-005', 'Verify Genuine News Analysis', 'Verify genuine text returns 94% confidence Genuine News.', 'PASSED', int((time.time() - start_time) * 1000))
        else:
            raise ValueError(f"Confidence value mismatch: got '{confidence_val}'")
    except Exception as e:
        log_test_result('TC-005', 'Verify Genuine News Analysis', 'Verify genuine text returns 94% confidence Genuine News.', 'FAILED', int((time.time() - start_time) * 1000), str(e))

    # Return back
    try:
        driver.back()
        time.sleep(1)
    except Exception:
        pass

    # TC-006: Navigate and Verify Trending Screen
    start_time = time.time()
    try:
        trending_btn = WebDriverWait(driver, 5).until(
            EC.presence_of_element_located((By.XPATH, '//android.widget.TextView[@text="📰 Trending News"]'))
        )
        trending_btn.click()
        
        # Check if list sources display
        WebDriverWait(driver, 5).until(
            EC.presence_of_element_located((By.XPATH, '//android.widget.TextView[contains(@text, "BBC") or contains(@text, "climate")]'))
        )
        log_test_result('TC-006', 'Verify Trending News List', 'Navigate to Trending and verify article sources load.', 'PASSED', int((time.time() - start_time) * 1000))
    except Exception as e:
        log_test_result('TC-006', 'Verify Trending News List', 'Navigate to Trending and verify article sources load.', 'FAILED', int((time.time() - start_time) * 1000), str(e))

    try:
        driver.back()
        time.sleep(1)
    except Exception:
        pass

    # TC-007: Navigate and Verify Dashboard Screen
    start_time = time.time()
    try:
        dashboard_btn = WebDriverWait(driver, 5).until(
            EC.presence_of_element_located((By.XPATH, '//android.widget.TextView[@text="📊 Dashboard"]'))
        )
        dashboard_btn.click()
        
        # Verify stats display
        WebDriverWait(driver, 5).until(
            EC.presence_of_element_located((By.XPATH, '//android.widget.TextView[contains(@text, "92%")]'))
        )
        log_test_result('TC-007', 'Verify Dashboard Statistics', 'Navigate to Dashboard and verify the 92% accuracy text.', 'PASSED', int((time.time() - start_time) * 1000))
    except Exception as e:
        log_test_result('TC-007', 'Verify Dashboard Statistics', 'Navigate to Dashboard and verify the 92% accuracy text.', 'FAILED', int((time.time() - start_time) * 1000), str(e))

    try:
        driver.back()
        time.sleep(1)
    except Exception:
        pass

    # TC-008: Navigate and Verify About Screen
    start_time = time.time()
    try:
        about_btn = WebDriverWait(driver, 5).until(
            EC.presence_of_element_located((By.XPATH, '//android.widget.TextView[@text="ℹ About"]'))
        )
        about_btn.click()
        
        # Verify educational purpose text
        WebDriverWait(driver, 5).until(
            EC.presence_of_element_located((By.XPATH, '//android.widget.TextView[contains(@text, "Educational Purpose")]'))
        )
        log_test_result('TC-008', 'Verify About Screen Metadata', 'Navigate to About and check educational purpose text.', 'PASSED', int((time.time() - start_time) * 1000))
    except Exception as e:
        log_test_result('TC-008', 'Verify About Screen Metadata', 'Navigate to About and check educational purpose text.', 'FAILED', int((time.time() - start_time) * 1000), str(e))

    if driver:
        print("Closing Appium mobile session...")
        driver.quit()

if __name__ == '__main__':
    try:
        run_appium_tests()
    except Exception as e:
        print(f"Appium E2E workflow error: {e}")
    finally:
        generate_excel_report()
        print("Mobile testing workflow completed.")
